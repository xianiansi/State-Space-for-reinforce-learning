import random
import numpy as np
from scipy.special import comb  # comb 组合数Cmn
from itertools import combinations
import copy
import matplotlib.pyplot as plt
import openpyxl
from mpl_toolkits.mplot3d import Axes3D


def Write_cell_terminate(score, teimanate, sheet):
    wb = openpyxl.load_workbook("result.xlsx")  # 生成一个已存在的wookbook对象
    wb1 = wb.worksheets[sheet]
    next_row = wb1.max_row + 1
    wb1.cell(next_row, 1, teimanate)
    for j1 in range(1,1+len(score)):
        wb1.cell(next_row, j1 + 1, score[j1-1])
    wb.save("result.xlsx")  # 保存

def uniformpoint(N, M):  # N:pop_size(总生成点个数),M:目标个数3
    H1 = 1
    while (comb(H1 + M - 1, M - 1) <= N):  # H1是每个方向上的方向数
        H1 = H1 + 1
    H1 = H1 - 1
    # H1+M-1中M-1个插板排列组合，
    W = np.array(list(combinations(range(H1 + M - 1), M - 1))) - np.tile(np.array(list(range(M - 1))),
                                                                         (int(comb(H1 + M - 1, M - 1)), 1))
    W = (np.hstack((W, H1 + np.zeros((W.shape[0], 1)))) - np.hstack(
        (np.zeros((W.shape[0], 1)), W))) / H1  # hstack水平拼接
    # 向量过于稀疏的情况
    if H1 < M:
        H2 = 0
        while (comb(H1 + M - 1, M - 1) + comb(H2 + M - 1, M - 1) <= N):
            H2 = H2 + 1
        H2 = H2 - 1
        if H2 > 0:
            W2 = np.array(list(combinations(range(H2 + M - 1), M - 1))) - np.tile(np.array(list(range(M - 1))),
                                                                                  (int(comb(H2 + M - 1, M - 1)), 1))
            W2 = (np.hstack((W2, H2 + np.zeros((W2.shape[0], 1)))) - np.hstack(
                (np.zeros((W2.shape[0], 1)), W2))) / H2
            W2 = W2 / 2 + 1 / (2 * M)
            W = np.vstack((W, W2))  # 垂直拼接
    W[W < 1e-6] = 1e-6  # 所有元素大于0
    N = W.shape[0]
    return W, N  # 一开始N是popsize，后面因为要生成平均向量，所以，种群个数改成能够形成组合数Cmn的那个N

def nonDominationSort(PopFun):
    nPop,nF = PopFun.shape[0],PopFun.shape[1]
    ranks = np.zeros(nPop, dtype=np.int32)
    nPs = np.zeros(nPop)  # 每个个体p被支配解的个数
    sPs = []  # 每个个体支配的解的集合，把索引放进去
    for i in range(nPop):
        iSet = []  # 解i的支配解集
        for j in range(nPop):
            if i == j:
                continue
            isDom1 = PopFun[i] <= PopFun[j]
            isDom2 = PopFun[i] < PopFun[j]
            # 是否支配该解-> i支配j
            if sum(isDom1) == nF and sum(isDom2) >= 1:
                iSet.append(j)
                # 是否被支配-> i被j支配
            if sum(~isDom2) == nF and sum(~isDom1) >= 1:
                nPs[i] += 1
        sPs.append(iSet)  # 添加i支配的解的索引
    r = 0  # 当前等级为 0， 等级越低越好
    indices = np.arange(nPop)
    while sum(nPs == 0) != 0:
        rIdices = indices[nPs == 0]  # 当前被支配数为0的索引
        ranks[rIdices] = r
        for rIdx in rIdices:
            iSet = sPs[rIdx]
            nPs[iSet] -= 1
        nPs[rIdices] = -1  # 当前等级的被支配数设置为负数
        r += 1
    return ranks
def isDominates(s1, s2):  # x是否支配y
    return (s1 <= s2).all() and (s1 < s2).any()

#指标
def IGD(popfun,PF,Zmin, Zmax):
    num_PF = PF.shape[0]
    popfun = (popfun-Zmin[0])/(Zmax[0]-Zmin[0])
    PF = (PF-Zmin[0])/(Zmax[0]-Zmin[0])
    distances = np.zeros(num_PF)
    for i, p_true in enumerate(PF):
        distances[i] = np.min(np.linalg.norm(popfun - p_true, axis=1))
    igd = np.sum(distances) / num_PF
    return igd
def GD(popfun, PF,Zmin, Zmax):
    num_pop = popfun.shape[0]
    popfun = (popfun - Zmin[0]) / (Zmax[0] - Zmin[0])
    PF = (PF - Zmin[0]) / (Zmax[0] - Zmin[0])
    distances = np.zeros(num_pop)
    for i, pop in enumerate(popfun):
        distances[i] = np.min(np.linalg.norm(pop - PF, axis=1))
    gd = np.sqrt(np.sum(distances)) / num_pop
    return gd
def HV(pf,Zmin, Zmax):
    # 输入Pareto Front的列表
    # pf = [(1, 1, 1), (2, 2, 2), (3, 4, 3), (4, 3, 4), (5, 5, 5)],ref_point = [(6, 6, 6)]
    ref_point = np.array([1,1])
    pf[:, 0] = (pf[:,0] - Zmin[0][0]) / (Zmax[0][0] - Zmin[0][0])
    pf[:, 1] = (pf[:, 1] - Zmin[0][1]) / (Zmax[0][1] - Zmin[0][1])
    # 定义超体积初始值
    hv = 0.0
    # 计算每个超体积贡献
    # ref_point_i = Zmin[0]
    for i in range(len(pf)):
        # 计算当前点到参考点的距离
        dist = (ref_point[0] - pf[i][0]) * (ref_point[1]-pf[i][1])
        # 如果距离为负数，则将其置为0
        if dist < 0:
            dist = 0
        # 累加超体积贡献
        hv += dist
    return hv
def Coverage(A, B):
    numB = 0  # B中被支配的个体
    for i in B:  # 对于集合B中的每一个个体
        for j in A:
            if isDominates(j, i):  # 至少被A中一个解支配
                numB += 1
                break
    ratio = numB / len(B)
    return ratio

color = [
    'red',
    'lightcoral',
    'orange',
    'darkseagreen',
    'green',
    'lightseagreen',
    'blue',
    'mediumorchid']


def ScatterPlot_2D(A,t):
    plt.scatter(A[:, 0], A[:, 1], marker="o", color=color[t//30], label=f"the {t}th iteration")

def ScatterPlot_3D(ax,A,t):
    ax.scatter(A[:, 0], A[:, 1], A[:, 2], marker="o", color=color[t // 30],label=f"the {t}th iteration")

# 初始算法，2目标
class MOEAD_2obj():
    # input: pop_size: 种群数量；XOVR：交叉概率；MUTR：变异概率；terminate: 迭代次数；archive：存档数目;T:邻域数量
    # instance: 算例对象
    def __init__(self,pop_size,XOVR, MUTR, terminate, archive,T,instance):
        # 算法输入参数
        self.pop_size = pop_size
        self.XOVR = XOVR
        self.MUTR = MUTR
        self.terminate = int(terminate)
        self.archive = archive  # 存档数目
        self.nObj = 2
        self.terminate = terminate
        self.T = T
        self.instance = instance

    # input： Chroms：进化的种群
    #         scene：决定解码的方式
    #         instance：参数对象
    def evolution(self,Chroms,scene):
        lamda, pop_size = uniformpoint(self.pop_size, self.nObj)  # Z是向量,N是向量个数(一般小于POP_SIZE)
        targs = self.instance.target_pop(Chroms, scene)
        Zmin = np.array(np.min(targs, 0)).reshape(1, self.nObj)  # 理想点（M个目标值 ）[1,1,1]三个目标的群体最小值
        B = self.look_neighbor(lamda)

        # 迭代过程
        ranks = nonDominationSort(targs)
        Pop_MOP = Chroms[ranks == 0]
        EPs = copy.deepcopy([list(Pop_MOP[i]) for i in range(len(Pop_MOP))])
        fig = plt.figure()
        for gen in range(self.terminate):
            print("第{name}次迭代".format(name=gen))
            for i in range(pop_size):
                ## 基因重组，从B(i)中随机选取两个序列k，l
                k = random.randint(0, self.T - 1)
                l = random.randint(0, self.T - 1)
                # flag有没有产生新解
                y = self.crossover(Chroms[B[i][k]], Chroms[B[i][l]])
                y = self.mutation(y)
                t_y, _ = self.instance.origin_target(y)

                ##更新z
                for j in range(len(Zmin[0])):
                    if t_y[j] < Zmin[0][j]:
                        Zmin[0][j] = t_y[j]
                ##更新领域解
                for j in range(len(B[i])):
                    gte_xi = self.Tchebycheff(Chroms[B[i][j]], lamda[B[i][j]], Zmin[0])
                    gte_y = self.Tchebycheff(y, lamda[B[i][j]], Zmin[0])
                    if (gte_y <= gte_xi):
                        Chroms[B[i][j]] = y


                ##更新外部存档
                ep = True  # 决定y是否放进EPs,True就放
                delete = []  # 装EPs个体
                for EP in EPs:  # EPs是所有支配个体，EPs外部存档
                    fun_EP,_ = self.instance.origin_target(EP)
                    if isDominates(fun_EP, t_y) or fun_EP.all() == t_y.all():  # EPs[k]支配y
                        ep = False
                        break
                    elif isDominates(t_y, fun_EP):  # 存在y支配EPs[k]
                        delete.append(EP)  # 准备删除EPs[k]
                if ep:  # 存在y支配EPs[k]或者没有任何支配关系
                    EPs.append(list(y))
                    for delete_i in delete[::-1]:
                        EPs.remove(delete_i)

            if gen % 10 == 0:
                print("%d gen has completed!\n" % gen)

            if gen % 30 == 0:
                targ_pop = self.instance.target_pop(Chroms, scene)
                ScatterPlot_2D(targ_pop, gen)
                print(Chroms)
                print('----------------------------')
                print(targ_pop)

            Zmin = np.array(np.min(targs, 0)).reshape(1, self.nObj)  # 求理想点
            Zmax = np.array(np.max(targs, 0)).reshape(1, self.nObj)  # 求负理想点
            targ_pop = self.instance.target_pop(Chroms, scene)
            c_min = Chroms[np.argmin(Chroms[:, 0])]
            Target_Ep = self.instance.target_pop(EPs, scene)
            score_GD = GD(targ_pop, Target_Ep, Zmin, Zmax)
            score_IGD = IGD(targ_pop, Target_Ep, Zmin, Zmax)
            score_HV = HV(Target_Ep, Zmin, Zmax)
            score = np.array([score_GD, score_IGD, score_HV])
            Write_cell_terminate(score, gen, sheet=0)

        plt.xlabel('Cmax')
        plt.ylabel('Qcost')
        plt.title("2-objectives Evolutionary Population Distribution")
        plt.grid(True)
        plt.legend()
        plt.show()


        Zmin = np.array(np.min(targs, 0)).reshape(1, self.nObj)  # 求理想点
        Zmax = np.array(np.max(targs, 0)).reshape(1, self.nObj)  # 求负理想点
        targ_pop = self.instance.target_pop(Chroms,scene)
        c_min = Chroms[np.argmin(Chroms[:, 0])]
        Target_Ep = self.instance.target_pop(EPs,scene)
        score_GD = GD(targ_pop, Target_Ep, Zmin, Zmax)
        score_IGD = IGD(targ_pop, Target_Ep, Zmin, Zmax)
        score_HV = HV(Target_Ep, Zmin, Zmax)
        score = np.array([score_GD, score_IGD, score_HV])
        ranks = nonDominationSort(targs)
        # pareto解集的目标
        Pop_MOP = Chroms[ranks == 0]
        targ_pop_ = self.instance.target_pop(Pop_MOP, scene)
        return Pop_MOP, targ_pop_, score



    def look_neighbor(self,lamda):
        B = []
        for i in range(len(lamda)):
            temp = []
            for j in range(len(lamda)):
                distance = np.sqrt((lamda[i][0] - lamda[j][0]) ** 2 +
                                   (lamda[i][1] - lamda[j][1]) ** 2)
                temp.append(distance)
            l = np.argsort(temp)
            B.append(l[:self.T])
        return B


    def crossover(self,s1, s2):
        # 两点交叉
        half_len = len(s1) // 2
        if random.random() < self.XOVR:
            temp = np.zeros(self.instance.J_num, dtype=int)  # 子代中每个工件已经有几道工序了
            offspring = np.zeros(len(s1), dtype=int)
            at1 = 0  # parent1指针
            at2 = 0  # parent2指针
            at = True  # 从哪个parent复制

            for i in range(half_len):
                while (offspring[i] == 0):  # 直到被赋值
                    if at:  # 从parent1取基因
                        j1 = s1[at1] - 1
                        if temp[j1] < self.instance.O_num[j1]:  # parent1对应的这个基因在子代中还没到达最大工序数
                            offspring[i] = s1[at1]  # 赋值
                            offspring[i + half_len] = s1[at1 + half_len]
                        at1 += 1  # 不管是否赋值，at1指针向后一格
                    else:  # 从parent2取基因
                        j2 = s2[at2] - 1
                        if temp[j2] < self.instance.O_num[j2]:
                            offspring[i] = s2[at2]
                            offspring[i + half_len] = s2[at2 + half_len]
                        at2 += 1
                    at = not at  # 逻辑取反
                temp[offspring[i] - 1] += 1


            # 检查可行性，修正染色体
            J_seq, O_seq, M_seq = self.instance.decode(offspring)
            for n in range(len(J_seq)):
                M_type = self.instance.OpType[J_seq[n] - 1][O_seq[n] - 1]  # OpType = [[2,2,4],[2,1,3,2],[3,1]]
                if M_seq[n] > self.instance.M_type_num[M_type - 1]:  ##随机选择一个机器
                    M_seq[n] = random.randint(1, self.instance.M_type_num[M_type - 1])
        else:
            offspring = random.choice((s1, s2))
        return offspring

    def mutation(self, offspring):
        J_seq, O_seq, M_seq = self.instance.decode(offspring)  # 解码后的
        half_len = len(offspring) // 2
        choice_pos_list = np.random.choice(half_len - 1, 10, replace=False)  # 选10个位置变异
        # print(J_seq, O_seq, M_seq)
        # if random.random() < self.MUTR:
        #     for pos in choice_pos_list:
        #         if J_seq[pos] != J_seq[pos + 1]:  # 前后交换不改变工序O_seq
        #             J_seq[pos], J_seq[pos + 1] = J_seq[pos + 1], J_seq[pos]
        #             M_seq[pos], M_seq[pos + 1] = M_seq[pos + 1], M_seq[pos]
        # print(J_seq, O_seq, M_seq)
        choice_pos_list = np.random.choice(half_len, 10, replace=False)  # 选10个位置变异
        if random.random() < self.MUTR:
            for pos in choice_pos_list:
                M_type = self.instance.OpType[J_seq[pos] - 1][O_seq[pos] - 1]
                M_seq[pos] = random.randint(1, self.instance.M_type_num[M_type - 1])
        # print(J_seq, O_seq, M_seq)
        offspring[:half_len] = J_seq
        offspring[half_len:] = M_seq
        return offspring

    def Tchebycheff(self, x, lamb, z):
        temp = []
        targ,_= self.instance.origin_target(x)
        for i in range(len(targ)):
            temp.append(np.abs(targ[i] - z[i]) * lamb[i])
        return np.max(temp)

    # 画图
    def plot_figure(self, C, PVal):
        # ---------------甘特图-------------
        cst = PVal[1][:] - PVal[0][:]  # 每道工序开始加工时刻
        cpt = PVal[0][:]  # 每道工序的加工时长
        cft = PVal[1][:]  # 每道工序的加工完成时刻
        cmn = PVal[2][:]  # 每道工序的加工机器
        osc = np.tile(C[0:int(len(C)/2)], 1)

        # ---------------甘特图-------------
        plt.figure()
        O_num_total = sum(self.instance.O_num)
        for i in range(O_num_total):
            if cft[i] != 0:
                plt.barh(y=cmn[i], width=cft[i] - cst[i], height=0.8, left=cst[i],
                         color=COLORS[osc[i] % LEN_COLORS], alpha=0.8, edgecolor="black")
                sf = r"$_{%s}$" % cst[i], r"$_{%s}$" % cft[i]
                x = cst[i], cft[i]
                # for j, k in enumerate(sf):  # 时间刻度
                #     plt.text(x=x[j], y=cmn[i], s=k,
                #              rotation="horizontal", va="top", ha="center")
                text = r"${%s}$" % (osc[i])  # 工件编号
                # text1 = r"${w%s}$" % (cwn[i])
                plt.text(x=cst[i] + 0.5 * cpt[i], y=cmn[i], s=text, c="black",
                         rotation="horizontal", va="center", ha="center")
                # plt.text(x=cst[i] + 0.8 * cpt[i], y=cmn[i], s=text1, c="black",
                #          rotation="horizontal", va="center", ha="center")
        plt.ylabel(r"Machine", fontsize=12, fontproperties="Arial")
        plt.xlabel(r"Makespan", fontsize=12, fontproperties="Arial")
        plt.title(r"Gantt Chart", fontsize=14, fontproperties="Arial")
        plt.show()


# 重调度，3目标
class MOEAD_3obj():
    # input: pop_size: 种群数量；XOVR：交叉概率；MUTR：变异概率；terminate: 迭代次数；archive：存档数目;T:邻域数量
    # instance: 算例对象
    def __init__(self,pop_size,XOVR, MUTR, terminate, archive, T, instance):
        # 算法输入参数
        self.pop_size = pop_size
        self.XOVR = XOVR
        self.MUTR = MUTR
        self.terminate = int(terminate)
        self.archive = archive  # 存档数目
        self.nObj = 3
        self.terminate = terminate
        self.T = T
        self.instance = instance

    # input： Chroms：进化的种群
    #         scene：决定解码的方式
    #         instance：参数对象
    def evolution(self,Chrom,J_temp, M_temp, pro_progress, re_progress, sq, sm):
        lamda, pop_size = uniformpoint(self.pop_size, self.nObj)  # Z是向量,N是向量个数(一般小于POP_SIZE)
        targs = self.instance.target_pops(Chrom, J_temp, M_temp, pro_progress, re_progress, sq, sm)
        Zmin = np.array(np.min(targs, 0)).reshape(1, self.nObj)  # 理想点（M个目标值 ）[1,1,1]三个目标的群体最小值
        B = self.look_neighbor(lamda)

        # 迭代过程
        ranks = nonDominationSort(targs)
        Pop_MOP = Chrom[ranks == 0]
        EPs = copy.deepcopy([list(Pop_MOP[i]) for i in range(len(Pop_MOP))])
        fig = plt.figure()
        # ax = Axes3D(fig)
        ax = fig.add_subplot(111, projection='3d')
        for gen in range(self.terminate):
            print("第{name}次迭代".format(name=gen))
            if gen % 30 == 0:
                targ_pop = self.instance.target_pops(Chrom,J_temp, M_temp,pro_progress, re_progress, sq, sm)
                # ScatterPlot_3D(ax, targ_pop, gen)
                print(Chrom)
                print('----------------------------')
                print(targ_pop)
                for targ in targ_pop:
                    Write_cell_terminate(targ, gen, sheet=0)

            for i in range(pop_size):
                ## 基因重组，从B(i)中随机选取两个序列k，l
                k = random.randint(0, self.T - 1)
                l = random.randint(0, self.T - 1)
                y = self.crossover(Chrom[B[i][k]], Chrom[B[i][l]], pro_progress, re_progress)
                y = self.mutation(y, pro_progress, re_progress)
                t_y = self.instance.target_pop(y,J_temp, M_temp, pro_progress, re_progress, sq, sm)

                ##更新z
                for j in range(len(Zmin[0])):
                    if t_y[j] < Zmin[0][j]:
                        Zmin[0][j] = t_y[j]
                ##更新领域解(只更新一个！！！！！)
                for j in range(len(B[i])):
                    # gte_xi = self.Tchebycheff(Chrom[B[i][j]], lamda[B[i][j]], Zmin[0],J_temp, M_temp, pro_progress, re_progress, sq, sm)
                    # gte_y = self.Tchebycheff(y, lamda[B[i][j]], Zmin[0], J_temp, M_temp, pro_progress, re_progress, sq, sm)
                    gte_xi = self.penalty_based_boundary(Chrom[B[i][j]], lamda[B[i][j]], Zmin[0], J_temp, M_temp, pro_progress,
                                              re_progress, sq, sm)
                    gte_y = self.penalty_based_boundary(y, lamda[B[i][j]], Zmin[0], J_temp, M_temp, pro_progress, re_progress, sq,
                                             sm)
                    if (gte_y <= gte_xi):
                        Chrom[B[i][j]] = y


                ##更新外部存档
                ep = True  # 决定y是否放进EPs,True就放
                delete = []  # 装EPs个体
                for EP in EPs:  # EPs是所有支配个体，EPs外部存档
                    fun_EP = self.instance.target_pop(EP,J_temp, M_temp, pro_progress, re_progress, sq, sm)
                    if isDominates(fun_EP, t_y) or fun_EP.all() == t_y.all():  # EPs[k]支配y
                        ep = False
                        break
                    elif isDominates(t_y, fun_EP):  # 存在y支配EPs[k]
                        delete.append(EP)  # 准备删除EPs[k]
                if ep:  # 存在y支配EPs[k]或者没有任何支配关系
                    EPs.append(list(y))
                    for delete_i in delete[::-1]:
                        EPs.remove(delete_i)

            if gen % 10 == 0:
                print("%d gen has completed!\n" % gen)



        # # ax.invert_yaxis()
        # print("aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa")
        # ax.view_init(elev=20, azim=-45)
        # ax.set_xlabel('Cmax')
        # ax.set_ylabel('Qcost')
        # ax.set_zlabel('robust')  # 给三个坐标轴注明坐标名称
        # plt.title("Three objectives ScatterPlot")
        # plt.grid(True)
        # plt.legend(loc='upper left', bbox_to_anchor=(1, 1))
        # plt.show()

        Zmin = np.array(np.min(targs, 0)).reshape(1, self.nObj)  # 求理想点
        Zmax = np.array(np.max(targs, 0)).reshape(1, self.nObj)  # 求负理想点
        targ_pop = self.instance.target_pops(Chrom, J_temp, M_temp, pro_progress, re_progress, sq, sm)
        Target_Ep = self.instance.target_pops(EPs,J_temp, M_temp, pro_progress, re_progress, sq, sm)
        score_GD = GD(targ_pop, Target_Ep, Zmin, Zmax)
        score_IGD = IGD(targ_pop, Target_Ep, Zmin, Zmax)
        score_HV = HV(Target_Ep, Zmin, Zmax)
        score = np.array([score_GD, score_IGD, score_HV])
        ranks = nonDominationSort(targs)
        Pop_MOP = Chrom[ranks == 0]
        return Pop_MOP, targ_pop, score


    def look_neighbor(self,lamda):
        B = []
        for i in range(len(lamda)):
            temp = []
            for j in range(len(lamda)):
                distance = np.sqrt((lamda[i][0] - lamda[j][0]) ** 2 +
                                   (lamda[i][1] - lamda[j][1]) ** 2)
                temp.append(distance)
            l = np.argsort(temp)
            B.append(l[:self.T])
        return B


    def crossover(self,s1, s2, pro_progress,re_progress):
        # 两点交叉
        half_len = len(s1) // 2
        if random.random() < self.XOVR:
            temp = np.zeros(self.instance.J_num, dtype=int)  # 子代中每个工件已经有几道工序了
            offspring = np.zeros(len(s1), dtype=int)
            at1 = 0  # parent1指针
            at2 = 0  # parent2指针
            at = True  # 从哪个parent复制

            for i in range(half_len):
                while (offspring[i] == 0):  # 直到被赋值
                    if at:  # 从parent1取基因
                        j1 = s1[at1]   # 工件号
                        if temp[j1 - 1] < self.instance.O_num[j1 - 1] - pro_progress[j1]:  # parent1对应的这个基因在子代中还没到达最大工序数
                            offspring[i] = s1[at1]  # 赋值
                            offspring[i + half_len] = s1[at1 + half_len]
                        at1 += 1  # 不管是否赋值，at1指针向后一格
                    else:  # 从parent2取基因
                        j2 = s2[at2]
                        if temp[j2 - 1] < self.instance.O_num[j2 -1] - pro_progress[j2]:
                            offspring[i] = s2[at2]
                            offspring[i + half_len] = s2[at2 + half_len]
                        at2 += 1
                    at = not at  # 逻辑取反
                temp[offspring[i] - 1] += 1


            # 检查可行性，修正染色体
            J_seq, O_seq, _, M_seq = self.instance.decode(offspring, pro_progress, re_progress)
            for n in range(len(J_seq)):
                M_type = self.instance.OpType[J_seq[n] - 1][O_seq[n] - 1]  # OpType = [[2,2,4],[2,1,3,2],[3,1]]
                if M_seq[n] > self.instance.M_type_num[M_type - 1]:  ##随机选择一个机器
                    M_seq[n] = random.randint(1, self.instance.M_type_num[M_type - 1])
        else:
            offspring = random.choice((s1, s2))
        return offspring

    def mutation(self, offspring, pro_progress, re_progress):
        J_seq, O_seq, _, M_seq = self.instance.decode(offspring, pro_progress, re_progress)  # 解码后的
        half_len = len(offspring) // 2
        choice_pos_list = np.random.choice(half_len, half_len//2, replace=False)  # 选一半的位置变异
        if random.random() < self.MUTR:
            for pos in choice_pos_list:
                M_type = self.instance.OpType[J_seq[pos] - 1][O_seq[pos] - 1]
                M_seq[pos] = random.randint(1, self.instance.M_type_num[M_type - 1])
        return offspring

    def Tchebycheff(self, x, lamb, z, J_temp, M_temp, pro_progress, re_progress, sq, sm):
        temp = []
        targ= self.instance.target_pop(x, J_temp, M_temp, pro_progress, re_progress, sq, sm)
        for i in range(len(targ)):
            temp.append(np.abs(targ[i] - z[i]) * lamb[i])
        return np.max(temp)

    def penalty_based_boundary(self, x, lamb, z,J_temp, M_temp, pro_progress, re_progress, sq, sm):
        targ = self.instance.target_pop(x, J_temp, M_temp, pro_progress, re_progress, sq, sm)
        d1 = np.linalg.norm(np.dot(targ - z, lamb)) / np.linalg.norm(lamb)
        d2 = np.linalg.norm(targ - (z + d1 * lamb / np.linalg.norm(lamb)))
        return d1 + 0.5 * d2

    # 画图
    def plot_figure(self, C, PVal):
        # ---------------甘特图-------------
        cst = PVal[1][:] - PVal[0][:]  # 每道工序开始加工时刻
        cpt = PVal[0][:]  # 每道工序的加工时长
        cft = PVal[1][:]  # 每道工序的加工完成时刻
        cmn = C[len(C)/2:]
        osc = np.tile(C[0:len(C)/2], 1)

        # ---------------甘特图-------------
        plt.figure()
        O_num_total = sum(self.instance.O_num)
        for i in range(O_num_total):
            if cft[i] != 0:
                plt.barh(y=cmn[i], width=cft[i] - cst[i], height=0.8, left=cst[i],
                         color=COLORS[osc[i] % LEN_COLORS], alpha=0.8, edgecolor="black")
                sf = r"$_{%s}$" % cst[i], r"$_{%s}$" % cft[i]
                x = cst[i], cft[i]
                # for j, k in enumerate(sf):  # 时间刻度
                #     plt.text(x=x[j], y=cmn[i], s=k,
                #              rotation="horizontal", va="top", ha="center")
                text = r"${%s}$" % (osc[i])  # 工件编号
                # text1 = r"${w%s}$" % (cwn[i])
                plt.text(x=cst[i] + 0.5 * cpt[i], y=cmn[i], s=text, c="black",
                         rotation="horizontal", va="center", ha="center")
                # plt.text(x=cst[i] + 0.8 * cpt[i], y=cmn[i], s=text1, c="black",
                #          rotation="horizontal", va="center", ha="center")
        plt.ylabel(r"Machine", fontsize=12, fontproperties="Arial")
        plt.xlabel(r"Makespan", fontsize=12, fontproperties="Arial")
        plt.title(r"Gantt Chart", fontsize=14, fontproperties="Arial")
        plt.show()








